from flask import Flask, render_template_string, redirect, url_for, request, session, flash, g, Response, jsonify
from flask_wtf.csrf import CSRFProtect
from dotenv import load_dotenv
import subprocess
import os
import signal
import threading
import sqlite3
import requests
import io
import csv
import time
from datetime import datetime, timedelta
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# --- CHARGEMENT DES SECRETS (.env) ---
load_dotenv()

app = Flask(__name__)

# --- CONFIGURATION SÉCURISÉE ---
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'YOUR_SECRET_KEY')

# Configuration des cookies de session (Sécurité)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=True,  # Mettre à True si vous êtes en HTTPS
    PERMANENT_SESSION_LIFETIME=timedelta(minutes=60)
)

# Protection CSRF (Cross-Site Request Forgery)
csrf = CSRFProtect(app)

# --- DÉCORATEUR POUR VÉRIFIER LA CLÉ API ---
def require_api_key(f):
    """Vérifie que la requête a la bonne clé secrète"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        api_key = request.headers.get('X-API-Key')
        if not api_key or api_key != app.secret_key:
            return jsonify({"error": "Clé API invalide"}), 401
        return f(*args, **kwargs)
    return decorated_function

GITHUB_CLIENT_ID = os.environ.get('GITHUB_CLIENT_ID')
GITHUB_CLIENT_SECRET = os.environ.get('GITHUB_CLIENT_SECRET')
GITHUB_REDIRECT_URI = os.environ.get('GITHUB_REDIRECT_URI', 'YOUR_WEBSITE/callback')
GITHUB_API_BASE_URL = "https://api.github.com"

# Clés Google reCAPTCHA
RECAPTCHA_SITE_KEY = os.environ.get('RECAPTCHA_SITE_KEY')
RECAPTCHA_SECRET_KEY = os.environ.get('RECAPTCHA_SECRET_KEY')

DB_PATH = 'DB_ACCESS_PATH'
BOTS_DIR = 'SCRIPTS_ACCESS_PATH'

# --- CREDENTIALS MANUELS ---
MANUAL_ADMIN_ID = "MANUAL_ADMIN"
MANUAL_ADMIN_USERNAME = "MANUAL_USERNAME"
MANUAL_LOGIN_ID = "MANUAL_LOGIN"
MANUAL_LOGIN_PASS = "MANUAL_PASSWORD"

# --- SYSTÈME DE TRADUCTION ---
TRANSLATIONS = {
    'fr': {
        'status_running': '🟢 EN COURS : ', 'status_stopped': '🔴 Arrêté', 'btn_stop': 'Arrêter BotJanus',
        'btn_start': 'DÉMARRER', 'script_running': 'Script en cours d\'exécution.', 'login_required': 'Connectez-vous pour lancer des scripts.',
        'locked_msg': '⛔ Lancement verrouillé par l\'administrateur.', 'console': 'Console', 'history': '📂 Historique',
        'my_account': '👤 Mon Compte', 'settings': '🛠 Paramètres', 'login_github': 'Connexion GitHub',
        'login_manual': 'Connexion Admin', 'logout': 'Se déconnecter', 'back': 'Retour', 'welcome': 'Bienvenue',
        'actions': 'Actions', 'banned': 'BANNI', 'ban': 'Bannir', 'unban': 'Débannir', 'update': 'Maj',
        'save': 'Enregistrer', 'users_roles': 'Utilisateurs & Rôles', 'system_settings': 'Paramètres Système',
        'security': 'Sécurité', 'lock_option': 'Verrouiller le lancement (admins seulement)', 'clean_logs': 'Nettoyer Logs',
        'login_title': 'Connexion Admin', 'username_ph': 'Identifiant', 'password_ph': 'Mot de passe',
        'connect_btn': 'Se connecter', 'lang_tag': 'Langue', 'role_tag': 'Rôle', 'days': 'jours',
        'error_auth': 'Erreur d\'authentification : Réservé à la connexion manuelle.', 'error_manual_login': 'Identifiants incorrects.'
    },
    'en': {
        'status_running': '🟢 RUNNING: ', 'status_stopped': '🔴 Stopped', 'btn_stop': 'Stop BotJanus',
        'btn_start': 'START', 'script_running': 'Script is currently running.', 'login_required': 'Please login to start scripts.',
        'locked_msg': '⛔ Launch locked by administrator.', 'console': 'Console', 'history': '📂 History',
        'my_account': '👤 My Account', 'settings': '🛠 Settings', 'login_github': 'GitHub Login',
        'login_manual': 'Admin Login', 'logout': 'Logout', 'back': 'Back', 'welcome': 'Welcome',
        'actions': 'Actions', 'banned': 'BANNED', 'ban': 'Ban', 'unban': 'Unban', 'update': 'Update',
        'save': 'Save', 'users_roles': 'Users & Roles', 'system_settings': 'System Settings', 'security': 'Security',
        'lock_option': 'Lock launching (Admins only)', 'clean_logs': 'Clean Logs', 'login_title': 'Admin Login',
        'username_ph': 'Username', 'password_ph': 'Password', 'connect_btn': 'Connect', 'lang_tag': 'Language',
        'role_tag': 'Role', 'days': 'days', 'error_auth': 'Auth Error: Manual login only.', 'error_manual_login': 'Incorrect credentials.'
    }
}

status = {
    "running": False, "process": None, "script_name": None,
    "live_output": [], "last_activity": None
}

ROLE_NONE = "None"
ROLE_COLLAB = "Collaborateur"
ROLE_ADMIN = "Admin"

# --- GESTION BASE DE DONNÉES ---

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DB_PATH)
        db.row_factory = sqlite3.Row
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    with app.app_context():
        try:
            db = get_db()
            db.execute('''CREATE TABLE IF NOT EXISTS logs (id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT, script TEXT, message TEXT)''')
            db.execute('''CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT)''')
            db.execute('''CREATE TABLE IF NOT EXISTS users (github_id TEXT PRIMARY KEY, username TEXT, avatar TEXT, role TEXT, is_banned INTEGER DEFAULT 0, lang TEXT DEFAULT 'fr', ban_reason TEXT)''')
            db.execute('''CREATE TABLE IF NOT EXISTS script_config (filename TEXT PRIMARY KEY, is_active INTEGER DEFAULT 1)''')
            db.execute('''CREATE TABLE IF NOT EXISTS schedules (id INTEGER PRIMARY KEY AUTOINCREMENT, script_name TEXT, frequency TEXT, time_value INTEGER, last_run TEXT, next_run TEXT, is_enabled INTEGER DEFAULT 1)''')
            db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('lock_launch', '0')")
            db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('captcha_enabled', '0')")
            db.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('auto_clean_days', '0')")
            db.commit()
        except Exception as e:
            print(f"Erreur DB Init: {e}")

init_db()

def log_to_db(script_name, message):
    try:
        conn = sqlite3.connect(DB_PATH)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute("INSERT INTO logs (date, script, message) VALUES (?, ?, ?)", (timestamp, script_name, message))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Erreur DB Insert: {e}")

# --- HELPER GESTION DYNAMIQUE DES SCRIPTS ---

def get_all_bots():
    if not os.path.exists(BOTS_DIR):
        try: os.makedirs(BOTS_DIR)
        except: pass

    # Création des fichiers par défaut si le dossier est vide lors de la première installation
    defaults = ["SCRIPT1.py", "SCRIPT2.py"]  # Remplacez par vos scripts par défaut
    for d in defaults:
        p = os.path.join(BOTS_DIR, d)
        if not os.path.exists(p):
            try:
                with open(p, "w", encoding="utf-8") as f:
                    f.write(f"# Script {d}\nprint('Initialisation du bot {d}...')\n")
            except: pass

    return sorted([f for f in os.listdir(BOTS_DIR) if f.endswith('.py')])

def get_script_status(filename):
    try:
        conn = sqlite3.connect(DB_PATH)
        res = conn.execute("SELECT is_active FROM script_config WHERE filename = ?", (filename,)).fetchone()
        conn.close()
        return res[0] if res is not None else 1
    except:
        return 1

# --- CORE LOGIQUE DE LANCEMENT ---

def launch_script_core(script_name, path, args=[], user_name="SYSTEM"):
    if status["running"]:
        return False

    cmd = ["python3", "-u", path] + args
    status["live_output"] = [f"--- Démarrage {script_name} par {user_name} ---"]
    if args:
        status["live_output"].append(f"Args: {' '.join(args)}")

    log_to_db("SYSTEM", f"Start {script_name} par {user_name}")

    try:
        process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, bufsize=1)
        threading.Thread(target=read_output, args=(process, script_name), daemon=True).start()
        status["running"], status["script_name"], status["process"], status["last_activity"] = True, script_name, process, datetime.now()
        return True
    except Exception as e:
        status["live_output"].append(f"Erreur de lancement : {str(e)}")
        return False

# --- DAEMON AUTOMATISÉ (PLANIFICATEUR) ---

def scheduler_loop():
    while True:
        try:
            time.sleep(30)
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row

            # --- NETTOYAGE AUTOMATIQUE DES LOGS ---
            res_clean = conn.execute("SELECT value FROM settings WHERE key='auto_clean_days'").fetchone()
            if res_clean and res_clean['value'] != '0':
                try:
                    days = int(res_clean['value'])
                    limit_date = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
                    conn.execute("DELETE FROM logs WHERE date < ?", (limit_date,))
                    conn.commit()
                except ValueError:
                    pass

            schedules = conn.execute("SELECT * FROM schedules WHERE is_enabled = 1").fetchall()

            for sched in schedules:
                next_run = sched['next_run']
                if next_run and next_run <= now_str:
                    if not status["running"]:
                        script_name = sched['script_name']
                        script_path = os.path.join(BOTS_DIR, script_name)
                        if os.path.exists(script_path):
                            # Lancement du traitement planifié
                            launch_script_core(script_name, script_path, args=[], user_name="Planificateur Auto")

                            # Calcul de l'occurrence suivante
                            freq = sched['frequency']
                            val = int(sched['time_value'] or 60)
                            last_run_time = datetime.now()
                            if freq == 'minutes':
                                next_run_time = last_run_time + timedelta(minutes=val)
                            elif freq == 'hours':
                                next_run_time = last_run_time + timedelta(hours=val)
                            elif freq == 'days':
                                next_run_time = last_run_time + timedelta(days=val)
                            else:
                                next_run_time = last_run_time + timedelta(days=1)

                            next_run_str = next_run_time.strftime("%Y-%m-%d %H:%M:%S")
                            conn.execute("UPDATE schedules SET last_run = ?, next_run = ? WHERE id = ?",
                                         (now_str, next_run_str, sched['id']))
                            conn.commit()
            conn.close()
        except Exception as e:
            print(f"Erreur du boucle du planificateur : {e}")

# Lancement du Thread Planificateur
threading.Thread(target=scheduler_loop, daemon=True).start()

# --- HELPER RECAPTCHA ---

def verify_recaptcha(response):
    payload = {'secret': RECAPTCHA_SECRET_KEY, 'response': response}
    try:
        r = requests.post('https://www.google.com/recaptcha/api/siteverify', data=payload)
        return r.json().get('success', False)
    except:
        return False

# --- PROTECTION GLOBALE (MIDDLEWARE) ---

@app.before_request
def check_security_gate():
    if request.endpoint in ['security_gate', 'verify_gate', 'static', 'callback', 'login_github'] or not request.endpoint:
        return

    db = sqlite3.connect(DB_PATH)
    res = db.execute("SELECT value FROM settings WHERE key='captcha_enabled'").fetchone()
    db.close()

    captcha_on = (res[0] == '1') if res else False

    if captcha_on and not session.get('captcha_passed'):
        return redirect(url_for('security_gate'))

# --- HELPER TRADUCTION ---
def get_text(key):
    lang = session.get('lang', 'fr')
    return TRANSLATIONS.get(lang, TRANSLATIONS['fr']).get(key, key)

@app.context_processor
def inject_globals():
    return dict(t=get_text, current_lang=session.get('lang', 'fr'))

# --- DECORATEURS ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash("Veuillez vous connecter.")
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def check_role(required_roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('index'))
            db = get_db()
            user = db.execute('SELECT * FROM users WHERE github_id = ?', (session['user_id'],)).fetchone()
            if not user or user['is_banned']:
                reason = user['ban_reason'] if user and user['ban_reason'] else "Non spécifiée"
                session.clear()
                flash(f"Compte banni. Raison : {reason}")
                return redirect(url_for('index'))
            if user['role'] not in required_roles and 'All' not in required_roles:
                flash("Droits insuffisants.")
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# ========================================
# ROUTES API DISCORD - BOT CONTROL
# ========================================

@app.route('/api/status', methods=['GET'])
@csrf.exempt
@require_api_key
def api_status():
    return jsonify({
        "running": status["running"],
        "script_name": status.get("script_name") or "Inactif",
        "last_activity": status["last_activity"].isoformat() if status["last_activity"] else None
    })

@app.route('/api/status_json')
def status_json():
    db = get_db()
    row = db.execute("SELECT value FROM settings WHERE key='lock_launch'").fetchone()
    locked = row['value'] if row else '0'
    user_role = ROLE_NONE
    if 'user_id' in session:
        u = db.execute("SELECT role FROM users WHERE github_id=?", (session['user_id'],)).fetchone()
        if u: user_role = u['role']
    return jsonify({
        "running": status["running"],
        "script_name": status.get("script_name") or "Inactif",
        "locked": locked,
        "role": user_role
    })

@app.route('/api/start', methods=['POST'])
@csrf.exempt
@require_api_key
def api_start():
    data = request.get_json()
    if not data:
        return jsonify({"error": "Aucune donnée reçue"}), 400

    choice = data.get('choice')
    discord_username = data.get('username', 'Discord')

    if not choice:
        return jsonify({"error": "Paramètre 'choice' manquant"}), 400

    if status["running"]:
        return jsonify({"error": "Un script est déjà en cours"}), 409

    p = os.path.join(BOTS_DIR, choice)
    if os.path.exists(p):
        cmd_args = []
        if choice == "portal.py" or choice == "Portail":
            cmd_args = ["--lang", data.get("arg_lang", "fr"), "--cat", data.get("arg_cat", ""), "--portal", data.get("arg_portal", "")]

        success = launch_script_core(choice, p, args=cmd_args, user_name=f"Discord ({discord_username})")
        if success:
            return jsonify({"success": True, "message": f"Script {choice} lancé."}), 200
    return jsonify({"error": "Fichier introuvable"}), 404

@app.route('/api/stop', methods=['POST'])
@csrf.exempt
@require_api_key
def api_stop():
    if not status["running"] or not status["process"]:
        return jsonify({"error": "Aucun script en cours"}), 400
    try:
        os.kill(status["process"].pid, signal.SIGTERM)
        status["running"], status["script_name"], status["process"] = False, None, None
        return jsonify({"success": True, "message": "Script arrêté."}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/logs', methods=['GET'])
@csrf.exempt
@require_api_key
def api_logs():
    limit = request.args.get('limit', 50, type=int)
    return jsonify({"logs": status["live_output"][-limit:]})

@app.route('/api/scripts', methods=['GET'])
@csrf.exempt
@require_api_key
def api_scripts():
    """Liste les scripts disponibles dans BOTS_DIR, pour le bot Discord."""
    scripts = []
    for filename in get_all_bots():
        scripts.append({
            "filename": filename,
            "is_active": bool(get_script_status(filename))
        })
    return jsonify({"scripts": scripts})

# --- CSS INTEGRANT LE DARK MODE AUTOMATIQUE ---
GLASS_CSS = """
<style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600&display=swap');
    * { box-sizing: border-box; }

    :root {
        --bg-gradient: linear-gradient(120deg, #0093E9 0%, #80D0C7 100%);
        --panel-bg: rgba(255, 255, 255, 0.15);
        --panel-border: rgba(255, 255, 255, 0.18);
        --panel-shadow: rgba(31, 38, 135, 0.20);
        --text-color: #fff;
        --input-bg: rgba(255, 255, 255, 0.1);
        --input-border: rgba(255, 255, 255, 0.3);
        --th-bg: rgba(0,0,0,0.2);
        --console-bg: rgba(0, 0, 0, 0.6);
        --console-border: rgba(255,255,255,0.1);
    }

    /* Configuration Mode Sombre Spécifique (Bleu Nuit & Noir Profond) */
    @media (prefers-color-scheme: dark) {
        :root {
            --bg-gradient: linear-gradient(135deg, #02040a 0%, #0b132b 50%, #000000 100%);
            --panel-bg: rgba(11, 19, 43, 0.55);
            --panel-border: rgba(255, 255, 255, 0.08);
            --panel-shadow: rgba(0, 0, 0, 0.6);
            --text-color: #f0f4f8;
            --input-bg: rgba(2, 4, 8, 0.7);
            --input-border: rgba(255, 255, 255, 0.15);
            --th-bg: rgba(5, 10, 20, 0.7);
            --console-bg: rgba(2, 4, 8, 0.95);
            --console-border: rgba(0, 180, 216, 0.3);
        }
    }

    body { font-family: 'Poppins', sans-serif; margin: 0; padding: 0; min-height: 100vh; background: var(--bg-gradient); background-size: 200% 200%; animation: gradientBG 15s ease infinite; color: var(--text-color); }
    @keyframes gradientBG { 0% { background-position: 0% 50%; } 50% { background-position: 100% 50%; } 100% { background-position: 0% 50%; } }
    .container { width: 95%; max-width: 900px; margin: 20px auto; padding-bottom: 80px; }
    .glass-panel { background: var(--panel-bg); box-shadow: 0 8px 32px 0 var(--panel-shadow); backdrop-filter: blur(12px); -webkit-backdrop-filter: blur(12px); border-radius: 16px; border: 1px solid var(--panel-border); padding: 25px; margin-bottom: 20px; text-align: center; }
    h1, h2, h3 { margin-top: 0; text-shadow: 0 2px 4px rgba(0,0,0,0.1); }
    .flex-row { display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 10px; }
    .flex-center { display: flex; justify-content: center; gap: 10px; flex-wrap: wrap; }
    .btn { padding: 10px 20px; font-size: 14px; font-weight: 600; border-radius: 50px; cursor: pointer; border: none; transition: 0.3s; text-decoration: none; display: inline-flex; justify-content: center; align-items: center; margin: 5px; box-shadow: 0 4px 15px rgba(0,0,0,0.2); white-space: nowrap; color: white; }
    .btn:hover { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(0,0,0,0.3); }
    .btn-github { background: #5865F2; } .btn-manual { background: #333; border: 1px solid #555; } .btn-action { background: linear-gradient(45deg, #11998e, #38ef7d); } .btn-danger { background: linear-gradient(45deg, #ff416c, #ff4b2b); } .btn-nav { background: rgba(255,255,255,0.3); border: 1px solid rgba(255,255,255,0.4); }
    input, select, textarea { padding: 10px; border-radius: 8px; border: 1px solid var(--input-border); background: var(--input-bg); color: var(--text-color); outline: none; }
    option { background: #111; color: white; }
    .table-responsive { overflow-x: auto; -webkit-overflow-scrolling: touch; }
    table { width: 100%; border-collapse: collapse; margin-top: 15px; min-width: 600px; }
    th, td { padding: 10px; text-align: left; border-bottom: 1px solid rgba(255,255,255,0.1); }
    th { background: var(--th-bg); }
    .console-window { background: var(--console-bg); border-radius: 12px; padding: 15px; height: 350px; overflow-y: scroll; text-align: left; font-family: 'Courier New', monospace; font-size: 13px; color: #0f0; border: 1px solid var(--console-border); }
    .alert { background: rgba(255, 200, 0, 0.3); padding: 10px; border-radius: 8px; margin-bottom: 15px; border: 1px solid orange; color: white; }
    .user-tag { display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 0.8em; font-weight: bold; }
    .role-Admin { background: #ff416c; } .role-Collaborateur { background: #38ef7d; color: #000; } .role-None { background: #ccc; color: #333; }
    .admin-nav-bar { display: flex; justify-content: center; gap: 10px; flex-wrap: wrap; margin-bottom: 20px; border-bottom: 1px solid rgba(255,255,255,0.1); padding-bottom: 15px; }
</style>
"""

# --- TEMPLATES HTML ---

GATE_HTML = """
<!DOCTYPE html>
<html>
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Sécurité - BotJanus</title>
  <script src="https://www.google.com/recaptcha/api.js" async defer></script>
  {{ glass_css|safe }}
</head>
<body>
  <div class="container" style="max-width:450px; padding-top:100px;">
    <div class="glass-panel">
        <h2>🛡️ Portail de Sécurité</h2>
        <form action="{{ url_for('verify_gate') }}" method="POST">
            <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
            <div class="g-recaptcha" data-sitekey="{{ site_key }}" style="display: inline-block; margin-bottom: 20px;"></div>
            <button type="submit" class="btn btn-action" style="width:100%;">Entrer sur le site</button>
        </form>
    </div>
  </div>
</body>
</html>
"""

DASHBOARD_HTML = """
<!DOCTYPE html>
<html>
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>BotJanus</title>
  {{ glass_css|safe }}
  <style>
    .modal-overlay { display: none; position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.7); z-index: 1000; justify-content: center; align-items: center; backdrop-filter: blur(5px); }
    .modal-content { background: rgba(25, 35, 50, 0.95); padding: 30px; border-radius: 16px; width: 90%; max-width: 400px; border: 1px solid rgba(255,255,255,0.2); text-align: left; }
    .form-group { margin-bottom: 15px; }
    .form-group label { display: block; margin-bottom: 5px; font-size: 0.9em; color: #ccc; }
    .form-group input, .form-group select { width: 100%; }
  </style>
</head>
<body>
  <div class="container">
    {% with messages = get_flashed_messages() %}
      {% if messages %}<div class="alert">{% for message in messages %}{{ message }}<br>{% endfor %}</div>{% endif %}
    {% endwith %}

    <div class="glass-panel flex-row">
        <div><h1>BotJanus</h1></div>
        <div>
            {% if session.get('user_id') %}
                <a href="{{ url_for('account') }}" class="btn btn-nav">{{ t('my_account') }}</a>
                {% if role == 'Admin' %}
                   <a href="{{ url_for('admin_users') }}" class="btn btn-danger">{{ t('settings') }}</a>
                {% endif %}
            {% else %}
                <a href="{{ url_for('login_github') }}" class="btn btn-github">{{ t('login_github') }}</a>
                <a href="{{ url_for('manual_login_page') }}" class="btn btn-manual">{{ t('login_manual') }}</a>
            {% endif %}
        </div>
    </div>

    <div class="glass-panel">
        <div style="margin-bottom: 10px;">Statut : <span style="font-weight:bold; color: {{ 'lightgreen' if running else 'salmon' }};">{{ t('status_running') + script_name if running else t('status_stopped') }}</span></div>
        {% if running %}
            {% if role in ['Collaborateur', 'Admin'] %}
                <form action="{{ url_for('stop_script') }}" method="post">
                    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
                    <button class="btn btn-danger" type="submit">{{ t('btn_stop') }}</button>
                </form>
            {% else %}
                <p><em>{{ t('script_running') }}</em></p>
            {% endif %}
        {% else %}
            {% if role in ['Collaborateur', 'Admin'] %}
                {% if locked == '1' and role != 'Admin' %}
                    <div style="background: rgba(255,0,0,0.2); padding: 15px; border-radius: 10px;">{{ t('locked_msg') }}</div>
                {% else %}
                    <form id="startForm" action="{{ url_for('start_script') }}" method="POST">
                        <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
                        <input type="hidden" name="arg_lang" id="hidden_lang">
                        <input type="hidden" name="arg_cat" id="hidden_cat">
                        <input type="hidden" name="arg_portal" id="hidden_portal">

                        <select id="scriptSelect" name="choice" style="min-width: 200px; margin-bottom: 10px;">
                            <optgroup label="Scripts Disponibles">
                                {% for s in available_scripts %}
                                    <option value="{{ s.filename }}">{{ s.filename }}{% if role == 'Admin' and s.is_active == 0 %} (Verrouillé aux Users){% endif %}</option>
                                {% endfor %}
                            </optgroup>
                        </select>
                        <button class="btn btn-action" type="button" onclick="handleStart()">{{ t('btn_start') }}</button>
                    </form>
                {% endif %}
            {% else %}
                <p><em>{{ t('login_required') }}</em></p>
            {% endif %}
        {% endif %}
    </div>

    <div class="glass-panel" style="text-align: left;">
        <div class="flex-row">
            <h3 style="margin:0;">{{ t('console') }}</h3>
            <a href="{{ url_for('history') }}" class="btn btn-nav">{{ t('history') }}</a>
        </div>
        <div class="console-window" id="logBox">
            {% for line in logs %}<div>{{ line }}</div>{% endfor %}
        </div>
    </div>
  </div>

  <div id="portalModal" class="modal-overlay">
      <div class="modal-content">
          <h3>⚙️ Configuration Portal Bot</h3>
          <div class="form-group">
              <label>Langue Vikidia</label>
              <select id="modal_lang">
                  <option value="fr">Français (fr)</option>
                  <option value="en">English (en)</option>
              </select>
          </div>
          <div class="form-group">
              <label>Nom de la Catégorie</label>
              <input type="text" id="modal_cat" placeholder="Ex: Histoire de France">
          </div>
          <div class="form-group">
              <label>Nom du Portail à ajouter</label>
              <input type="text" id="modal_portal" placeholder="Ex: France">
          </div>
          <div style="text-align:right; margin-top:20px;">
              <button class="btn btn-nav" onclick="closeModal()">Annuler</button>
              <button class="btn btn-action" onclick="confirmPortalLaunch()">Lancer</button>
          </div>
      </div>
  </div>

  <script>
    var logBox = document.getElementById("logBox");
    if(logBox) {
        logBox.scrollTop = logBox.scrollHeight;
        setInterval(function(){
            fetch("/api/live_logs").then(r => r.text()).then(data => {
                let isScrolled = logBox.scrollHeight - logBox.clientHeight <= logBox.scrollTop + 50;
                logBox.innerHTML = data;
                if(isScrolled) logBox.scrollTop = logBox.scrollHeight;
            });
        }, 2000);
    }

    // SYSTEM DE NOTIFICATION PUSH CE CÔTÉ CLIENT
    if (window.Notification && Notification.permission === "default") {
        Notification.requestPermission();
    }

    let wasRunning = {{ 'true' if running else 'false' }};
    let activeScriptName = "{{ script_name }}";

    setInterval(function(){
        fetch("/api/status_json").then(r => r.json()).then(res => {
            if (wasRunning && !res.running) {
                if (window.Notification && Notification.permission === "granted") {
                    new Notification("🤖 BotJanus - Script Terminé", {
                        body: "Le traitement du script '" + activeScriptName + "' s'est achevé.",
                        icon: "https://ui-avatars.com/api/?name=Bot+Janus&background=0093E9&color=fff"
                    });
                }
            }
            wasRunning = res.running;
            activeScriptName = res.script_name;
        });
    }, 3000);

    function handleStart() {
        var choice = document.getElementById('scriptSelect').value;
        if (choice.includes('portal.py') || choice === 'Portail') {
            document.getElementById('portalModal').style.display = 'flex';
        } else {
            document.getElementById('startForm').submit();
        }
    }
    function closeModal() { document.getElementById('portalModal').style.display = 'none'; }
    function confirmPortalLaunch() {
        document.getElementById('hidden_lang').value = document.getElementById('modal_lang').value;
        document.getElementById('hidden_cat').value = document.getElementById('modal_cat').value;
        document.getElementById('hidden_portal').value = document.getElementById('modal_portal').value;
        document.getElementById('startForm').submit();
        closeModal();
    }
  </script>
</body>
</html>
"""

ADMIN_SCRIPTS_HTML = """
<!DOCTYPE html>
<html>
<head>
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Scripts - Administration</title>
  {{ glass_css|safe }}
</head>
<body>
  <div class="container">
    {% with messages = get_flashed_messages() %}
      {% if messages %}<div class="alert">{% for message in messages %}{{ message }}<br>{% endfor %}</div>{% endif %}
    {% endwith %}

    <div class="glass-panel">
        <h2>🛠️ Espace Administration</h2>
        <div class="admin-nav-bar">
            <a href="{{ url_for('admin_users') }}" class="btn btn-nav">Utilisateurs</a>
            <a href="{{ url_for('admin_scripts') }}" class="btn btn-action">Scripts</a>
            <a href="{{ url_for('admin_schedules') }}" class="btn btn-nav">Planification</a>
            <a href="{{ url_for('admin_stats') }}" class="btn btn-nav">Statistiques</a>
            <a href="{{ url_for('settings') }}" class="btn btn-nav">Système</a>
            <a href="{{ url_for('index') }}" class="btn btn-nav">← Dashboard</a>
        </div>
    </div>

    <div class="glass-panel" style="text-align: left;">
        <h3>📁 Créer un script (.py)</h3>
        <form action="{{ url_for('admin_create_script') }}" method="POST" style="display: flex; gap: 10px;">
            <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
            <input type="text" name="filename" placeholder="Ex: mon_bot.py" required style="flex: 1;">
            <button type="submit" class="btn btn-action" style="margin:0;">Créer</button>
        </form>
    </div>

    <div class="glass-panel" style="text-align: left;">
        <h3>📜 Liste des Scripts (/bots)</h3>
        <div class="table-responsive">
            <table>
                <tr><th>Nom du Fichier</th><th>Accès Collaborateurs</th><th>Actions</th></tr>
                {% for script in scripts %}
                <tr>
                    <td><b>{{ script }}</b></td>
                    <td>
                        {% set is_active = configs.get(script, 1) %}
                        <span class="user-tag" style="background: {{ '#38ef7d' if is_active == 1 else '#ff416c' }}; color: {{ '#000' if is_active == 1 else '#fff' }};">
                            {{ 'Disponible' if is_active == 1 else 'Masqué' }}
                        </span>
                    </td>
                    <td>
                        <div style="display: flex; gap: 5px;">
                            <form action="{{ url_for('admin_toggle_script') }}" method="POST" style="margin:0;">
                                <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
                                <input type="hidden" name="filename" value="{{ script }}">
                                <input type="hidden" name="current_val" value="{{ is_active }}">
                                <button type="submit" class="btn btn-nav" style="padding:5px 10px; font-size:0.85em;">Changer Droits</button>
                            </form>
                            <a href="{{ url_for('admin_edit_script_page', filename=script) }}" class="btn btn-manual" style="padding:5px 10px; font-size:0.85em;">✏️ Éditer</a>
                            <form action="{{ url_for('admin_run_script_direct') }}" method="POST" style="margin:0;">
                                <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
                                <input type="hidden" name="filename" value="{{ script }}">
                                <button type="submit" class="btn btn-action" style="padding:5px 10px; font-size:0.85em;">⚡ Run Direct</button>
                            </form>
                        </div>
                    </td>
                </tr>
                {% endfor %}
            </table>
        </div>
    </div>
  </div>
</body>
</html>
"""

SCRIPT_EDITOR_HTML = """
<!DOCTYPE html>
<html>
<head>
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Éditeur - {{ filename }}</title>
  {{ glass_css|safe }}
</head>
<body>
  <div class="container">
    <div class="glass-panel">
        <div class="flex-row">
            <h2>✏️ Éditeur : {{ filename }}</h2>
            <a href="{{ url_for('admin_scripts') }}" class="btn btn-nav">Retour</a>
        </div>
        <form action="{{ url_for('admin_save_script') }}" method="POST" style="text-align: left; margin-top:15px;">
            <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
            <input type="hidden" name="filename" value="{{ filename }}"/>
            <textarea name="content" style="width: 100%; height: 450px; font-family: monospace; font-size: 13px; background: #070c14; color: #00ff66; padding: 10px; border-radius: 8px;">{{ content }}</textarea>
            <button type="submit" class="btn btn-action" style="width:100%; margin-top:10px;">Sauvegarder le Script</button>
        </form>
    </div>
  </div>
</body>
</html>
"""

SCHEDULES_HTML = """
<!DOCTYPE html>
<html>
<head>
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Planification - Admin</title>
  {{ glass_css|safe }}
</head>
<body>
  <div class="container">
    <div class="glass-panel">
        <h2>⏳ Automatisation Planifiée</h2>
        <div class="admin-nav-bar">
            <a href="{{ url_for('admin_users') }}" class="btn btn-nav">Utilisateurs</a>
            <a href="{{ url_for('admin_scripts') }}" class="btn btn-nav">Scripts</a>
            <a href="{{ url_for('admin_schedules') }}" class="btn btn-action">Planification</a>
            <a href="{{ url_for('admin_stats') }}" class="btn btn-nav">Statistiques</a>
            <a href="{{ url_for('settings') }}" class="btn btn-nav">Système</a>
            <a href="{{ url_for('index') }}" class="btn btn-nav">← Dashboard</a>
        </div>
    </div>

    <div class="glass-panel" style="text-align: left;">
        <h3>⏰ Ajouter une tâche</h3>
        <form action="{{ url_for('admin_schedules') }}" method="POST" style="display: flex; gap: 10px; flex-wrap: wrap;">
            <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
            <input type="hidden" name="action" value="add"/>
            <select name="script_name">
                {% for s in scripts %}<option value="{{ s }}">{{ s }}</option>{% endfor %}
            </select>
            <select name="frequency">
                <option value="minutes">Minutes</option>
                <option value="hours">Heures</option>
                <option value="days">Jours</option>
            </select>
            <input type="number" name="time_value" value="30" style="width:80px;">
            <button type="submit" class="btn btn-action" style="margin:0;">Créer Répétition</button>
        </form>
    </div>

    <div class="glass-panel" style="text-align: left;">
        <h3>📋 Planifications actives</h3>
        <div class="table-responsive">
            <table>
                <tr><th>Script</th><th>Intervalle</th><th>Dernier Run</th><th>Prochain Run</th><th>Action</th></tr>
                {% for s in schedules %}
                <tr>
                    <td><b>{{ s.script_name }}</b></td>
                    <td>Toutes les {{ s.time_value }} {{ s.frequency }}</td>
                    <td>{{ s.last_run }}</td>
                    <td style="color:#38ef7d;">{{ s.next_run }}</td>
                    <td>
                        <form action="{{ url_for('admin_schedules') }}" method="POST" style="margin:0;">
                            <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
                            <input type="hidden" name="action" value="delete"/>
                            <input type="hidden" name="id" value="{{ s.id }}"/>
                            <button type="submit" class="btn btn-danger" style="padding:5px;">Supprimer</button>
                        </form>
                    </td>
                </tr>
                {% endfor %}
            </table>
        </div>
    </div>
  </div>
</body>
</html>
"""

STATS_HTML = """
<!DOCTYPE html>
<html>
<head>
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Statistiques - Admin</title>
  {{ glass_css|safe }}
</head>
<body>
  <div class="container">
    <div class="glass-panel">
        <h2>📊 Rapports & Statistiques</h2>
        <div class="admin-nav-bar">
            <a href="{{ url_for('admin_users') }}" class="btn btn-nav">Utilisateurs</a>
            <a href="{{ url_for('admin_scripts') }}" class="btn btn-nav">Scripts</a>
            <a href="{{ url_for('admin_schedules') }}" class="btn btn-nav">Planification</a>
            <a href="{{ url_for('admin_stats') }}" class="btn btn-action">Statistiques</a>
            <a href="{{ url_for('settings') }}" class="btn btn-nav">Système</a>
            <a href="{{ url_for('index') }}" class="btn btn-nav">← Dashboard</a>
        </div>
    </div>

    <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px;">
        <div class="glass-panel"><h3>Membres inscrits</h3><p style="font-size:2em; color:#38ef7d;">{{ total_users }}</p></div>
        <div class="glass-panel"><h3>Volume Logs</h3><p style="font-size:2em; color:#00b4d8;">{{ total_logs }}</p></div>
    </div>

    <div class="glass-panel" style="text-align: left;">
        <h3>📈 Activité par Module (Lignes de logs)</h3>
        <table>
            <tr><th>Nom du Module</th><th>Volume d'activité</th></tr>
            {% for c in script_counts %}
            <tr><td>{{ c.script }}</td><td><b>{{ c.cnt }}</b> lignes</td></tr>
            {% endfor %}
        </table>
    </div>
  </div>
</body>
</html>
"""

LOGIN_MANUAL_HTML = """
<!DOCTYPE html>
<html>
<head>
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{{ t('login_title') }}</title>
  {{ glass_css|safe }}
</head>
<body>
<div class="container" style="max-width:400px; padding-top:100px;">
    <div class="glass-panel">
        <h2>{{ t('login_title') }}</h2>
        <form action="{{ url_for('manual_login_post') }}" method="POST">
            <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
            <input type="text" name="username" placeholder="{{ t('username_ph') }}" style="width:100%; margin-bottom:10px;"><br>
            <input type="password" name="password" placeholder="{{ t('password_ph') }}" style="width:100%; margin-bottom:10px;"><br>
            <button class="btn btn-action" type="submit" style="width:100%;">{{ t('connect_btn') }}</button>
        </form>
        <br><a href="{{ url_for('index') }}" class="btn btn-nav" style="width:100%;">{{ t('back') }}</a>
    </div>
</div>
</body>
</html>
"""

ACCOUNT_HTML = """
<!DOCTYPE html>
<html>
<head>
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{{ t('my_account') }}</title>
  {{ glass_css|safe }}
</head>
<body>
  <div class="container" style="max-width: 500px;">
    <div class="glass-panel">
        <h2>{{ t('my_account') }}</h2>
        <img src="{{ user.avatar }}" style="width:80px; border-radius:50%; border: 2px solid white; margin-bottom:10px;">
        <h3>{{ user.username }}</h3>
        <p>{{ t('role_tag') }} : <span class="user-tag role-{{ user.role }}">{{ user.role }}</span></p>
        <div style="margin: 20px 0; padding: 15px; background: rgba(255,255,255,0.05); border-radius: 12px;">
            <label for="lang-select" style="margin-right: 10px; font-weight: bold;">{{ t('lang_tag') }} :</label>
            <select id="lang-select" onchange="window.location.href = '{{ url_for('set_language', code='') }}' + this.value">
                <option value="fr" {{ 'selected' if current_lang == 'fr' else '' }}>Français</option>
                <option value="en" {{ 'selected' if current_lang == 'en' else '' }}>English</option>
            </select>
        </div>
        <hr style="border-color: rgba(255,255,255,0.2); margin-bottom: 20px;">
        <a href="{{ url_for('logout') }}" class="btn btn-danger">{{ t('logout') }}</a>
        <a href="{{ url_for('index') }}" class="btn btn-nav">{{ t('back') }}</a>
    </div>
  </div>
</body>
</html>
"""

ADMIN_USERS_HTML = """
<!DOCTYPE html>
<html>
<head>
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{{ t('users_roles') }}</title>
  {{ glass_css|safe }}
</head>
<body>
  <div class="container">
    <div class="glass-panel">
        <h2>{{ t('settings') }}</h2>
        <div class="admin-nav-bar">
            <a href="{{ url_for('admin_users') }}" class="btn btn-action">Utilisateurs</a>
            <a href="{{ url_for('admin_scripts') }}" class="btn btn-nav">Scripts</a>
            <a href="{{ url_for('admin_schedules') }}" class="btn btn-nav">Planification</a>
            <a href="{{ url_for('admin_stats') }}" class="btn btn-nav">Statistiques</a>
            <a href="{{ url_for('settings') }}" class="btn btn-nav">Système</a>
            <a href="{{ url_for('index') }}" class="btn btn-nav">← Dashboard</a>
        </div>
    </div>
    <div class="glass-panel" style="text-align:left;">
        <h3>✅ {{ t('users_roles') }}</h3>
        <div class="table-responsive">
            <table>
                <tr><th>User</th><th>{{ t('role_tag') }}</th><th>{{ t('actions') }}</th></tr>
                {% for u in users if not u.is_banned %}
                <tr>
                    <td><b>{{ u.username }}</b></td>
                    <td><span class="user-tag role-{{ u.role }}">{{ u.role }}</span></td>
                    <td>
                        <form action="{{ url_for('admin_update_user') }}" method="POST" id="form-{{ u.github_id }}" style="display:flex; gap:5px;">
                            <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
                            <input type="hidden" name="github_id" value="{{ u.github_id }}">
                            <input type="hidden" name="reason" id="reason-{{ u.github_id }}" value="">
                            <input type="hidden" name="action" id="action-{{ u.github_id }}" value="update">
                            <select name="new_role" style="padding:5px;">
                                <option value="None" {{ 'selected' if u.role == 'None' else '' }}>None</option>
                                <option value="Collaborateur" {{ 'selected' if u.role == 'Collaborateur' else '' }}>Collaborateur</option>
                                <option value="Admin" {{ 'selected' if u.role == 'Admin' else '' }}>Admin</option>
                            </select>
                            <button type="submit" class="btn btn-nav" style="padding:5px 10px; margin:0; font-size:0.8em;">{{ t('update') }}</button>
                            <button type="button" onclick="confirmBan('{{ u.github_id }}', '{{ u.username }}')" class="btn btn-danger" style="padding:5px 10px; margin:0; font-size:0.8em;">{{ t('ban') }}</button>
                        </form>
                    </td>
                </tr>
                {% endfor %}
            </table>
        </div>
    </div>
  </div>
  <script>
    function confirmBan(userId, username) {
        let reason = prompt("Raison du bannissement pour " + username + " ?");
        if (reason) {
            document.getElementById('reason-' + userId).value = reason;
            document.getElementById('action-' + userId).value = 'ban';
            document.getElementById('form-' + userId).submit();
        }
    }
  </script>
</body>
</html>
"""

SETTINGS_HTML = """
<!DOCTYPE html>
<html>
<head>
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{{ t('system_settings') }}</title>
  {{ glass_css|safe }}
</head>
<body>
  <div class="container" style="max-width: 600px;">
    {% with messages = get_flashed_messages() %}
      {% if messages %}<div class="alert">{% for message in messages %}{{ message }}<br>{% endfor %}</div>{% endif %}
    {% endwith %}
    <div class="glass-panel">
        <h2>{{ t('system_settings') }}</h2>
        <div class="admin-nav-bar">
            <a href="{{ url_for('admin_users') }}" class="btn btn-nav">Utilisateurs</a>
            <a href="{{ url_for('admin_scripts') }}" class="btn btn-nav">Scripts</a>
            <a href="{{ url_for('admin_schedules') }}" class="btn btn-nav">Planification</a>
            <a href="{{ url_for('admin_stats') }}" class="btn btn-nav">Statistiques</a>
            <a href="{{ url_for('settings') }}" class="btn btn-action">Système</a>
            <a href="{{ url_for('index') }}" class="btn btn-nav">← Dashboard</a>
        </div>
    </div>

    <div class="glass-panel" style="text-align: left;">
        <h3>💾 Sauvegardes de Données (Format Excel)</h3>
        <p style="font-size:0.85em; opacity:0.8;">Exportez les profils ou effectuez une restauration en cas d'anomalie système.</p>
        <a href="/admin/backup/export" class="btn btn-action" style="margin-left:0;">📥 Exporter la base (CSV)</a>
        <hr style="border-color:rgba(255,255,255,0.1); margin:15px 0;">
        <form action="/admin/backup/import" method="POST" enctype="multipart/form-data">
            <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
            <input type="file" name="backup_file" accept=".csv" required><br><br>
            <button type="submit" class="btn btn-danger" style="margin:0;">📤 Restaurer le fichier</button>
        </form>
    </div>

    <div class="glass-panel" style="text-align: left;">
        <h3>🧹 Gestion et Nettoyage des Logs</h3>

        <form action="/admin/clean_logs_manual" method="POST" style="margin-bottom: 20px;">
            <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
            <p style="font-size:0.85em; opacity:0.8;"><b>Nettoyage manuel :</b> Supprimer définitivement les logs plus anciens que X jours.</p>
            <div style="display: flex; gap: 10px; align-items: center;">
                <input type="number" name="days" min="0" placeholder="Ex: 7" required style="width: 100px;">
                <span>jours</span>
                <button type="submit" class="btn btn-danger" style="margin:0;">Supprimer manuellement</button>
            </div>
        </form>
    </div>

    <div class="glass-panel" style="text-align: left;">
        <h3>{{ t('security') }}</h3>
        <form action="{{ url_for('update_settings') }}" method="POST">
            <input type="hidden" name="csrf_token" value="{{ csrf_token() }}"/>
            <label style="display: flex; align-items: center; margin-bottom:15px;">
                <input type="checkbox" name="lock_launch" value="1" {% if locked == '1' %}checked{% endif %} style="width: 20px; height: 20px; margin-right: 10px;">
                <span>{{ t('lock_option') }}</span>
            </label>
            <label style="display: flex; align-items: center; margin-bottom:15px;">
                <input type="checkbox" name="captcha_enabled" value="1" {% if captcha_enabled == '1' %}checked{% endif %} style="width: 20px; height: 20px; margin-right: 10px;">
                <span>Activer le portail reCAPTCHA Google</span>
            </label>

            <hr style="border-color:rgba(255,255,255,0.1); margin:15px 0;">
            <p style="font-size:0.85em; opacity:0.8; margin-bottom: 10px;"><b>Nettoyage automatique :</b> Conserver uniquement les logs des X derniers jours (0 pour désactiver).</p>
            <div style="display: flex; gap: 10px; align-items: center; margin-bottom: 15px;">
                <input type="number" name="auto_clean_days" min="0" value="{{ auto_clean_days }}" required style="width: 100px;">
                <span>jours</span>
            </div>

            <button class="btn btn-action" type="submit">{{ t('save') }}</button>
        </form>
    </div>
  </div>
</body>
</html>
"""

HISTORY_HTML = """
<!DOCTYPE html>
<html>
<head>
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{{ t('history') }}</title>
  {{ glass_css|safe }}
</head>
<body>
  <div class="container" style="max-width: 1100px;">
    <div class="glass-panel" style="text-align: left;">
        <div class="flex-row">
            <h2 style="margin:0;">📂 Logs</h2>
            <div style="display:flex; gap:8px; flex-wrap:wrap; align-items:center;">
                <a href="{{ url_for('export_logs_excel') }}{% if request.args.get('search') %}?search={{ request.args.get('search') }}{% endif %}" class="btn btn-action" style="margin:0; padding:8px 16px; font-size:0.85em;">📊 Export Excel</a>
                <a href="{{ url_for('export_logs_csv') }}{% if request.args.get('search') %}?search={{ request.args.get('search') }}{% endif %}" class="btn btn-nav" style="margin:0; padding:8px 16px; font-size:0.85em;">📄 Export CSV</a>
                <a href="{{ url_for('index') }}" class="btn btn-nav" style="margin:0;">← {{ t('back') }}</a>
            </div>
        </div>
        <form id="filter-form" action="{{ url_for('history') }}" method="GET" style="margin-top:15px;">
            <input type="text" name="search" placeholder="Rechercher un message..." value="{{ request.args.get('search', '') }}">
            <button type="submit" class="btn btn-nav">🔍</button>
        </form>
        <p style="font-size:0.8em; opacity:0.6; margin:8px 0 0 0;">{{ total_count }} entrée(s) trouvée(s) — affichage des 500 plus récentes</p>
        <div class="table-responsive">
            <table>
                <tr><th>Date</th><th>Script</th><th>Message</th></tr>
                {% for row in rows %}
                <tr>
                    <td style="color: #a8dadc; white-space:nowrap;">{{ row[1] }}</td>
                    <td>{{ row[2] }}</td>
                    <td>{{ row[3] }}</td>
                </tr>
                {% endfor %}
            </table>
        </div>
    </div>
  </div>
</body>
</html>
"""

# --- INCLUSION DES SYSTEMES DE BACKUP / RESTAURATION ---

@app.route("/admin/backup/export")
@check_role([ROLE_ADMIN])
def backup_export():
    db = get_db()
    users = db.execute("SELECT * FROM users").fetchall()
    si = io.StringIO()
    cw = csv.writer(si, delimiter=';')
    cw.writerow(['github_id', 'username', 'avatar', 'role', 'is_banned', 'lang', 'ban_reason'])
    for u in users:
        cw.writerow([u['github_id'], u['username'], u['avatar'], u['role'], u['is_banned'], u['lang'], u['ban_reason']])
    return Response(
        si.getvalue(),
        mimetype="text/csv",
        headers={"Content-disposition": f"attachment; filename=backup_users_{datetime.now().strftime('%Y%m%d')}.csv"}
    )

@app.route("/admin/backup/import", methods=["POST"])
@check_role([ROLE_ADMIN])
def backup_import():
    if 'backup_file' not in request.files:
        flash("Fichier manquant.")
        return redirect(url_for('settings'))
    file = request.files['backup_file']
    if file.filename == '':
        return redirect(url_for('settings'))
    try:
        stream = io.StringIO(file.stream.read().decode("UTF-8"), newline=None)
        reader = csv.reader(stream, delimiter=';')
        next(reader)  # Sauter l'entête
        db = get_db()
        count = 0
        for row in reader:
            if len(row) >= 6:
                db.execute("""INSERT OR REPLACE INTO users (github_id, username, avatar, role, is_banned, lang, ban_reason)
                              VALUES (?, ?, ?, ?, ?, ?, ?)""",
                           (row[0], row[1], row[2], row[3], int(row[4]), row[5], row[6] if len(row) > 6 else ""))
                count += 1
        db.commit()
        flash(f"Restauration terminée : {count} profils importés.")
    except Exception as e:
        flash(f"Erreur d'importation : {str(e)}")
    return redirect(url_for('settings'))

# --- INTERFACES SCRIPTS ET AUTOMATISATION POUR L'ADMINISTRATEUR ---

@app.route("/admin/scripts")
@check_role([ROLE_ADMIN])
def admin_scripts():
    scripts = get_all_bots()
    configs = {}
    db = get_db()
    rows = db.execute("SELECT filename, is_active FROM script_config").fetchall()
    for r in rows:
        configs[r['filename']] = r['is_active']
    return render_template_string(ADMIN_SCRIPTS_HTML, scripts=scripts, configs=configs, glass_css=GLASS_CSS)

@app.route("/admin/scripts/create", methods=["POST"])
@check_role([ROLE_ADMIN])
def admin_create_script():
    filename = request.form.get("filename", "").strip()
    if not filename.endswith(".py"): filename += ".py"
    if len(filename) > 4:
        p = os.path.join(BOTS_DIR, filename)
        if not os.path.exists(p):
            with open(p, "w", encoding="utf-8") as f:
                f.write("#!/usr/bin/env python3\nprint('Nouveau script bot.')\n")
            flash("Fichier script créé.")
    return redirect(url_for('admin_scripts'))

@app.route("/admin/scripts/edit/<filename>")
@check_role([ROLE_ADMIN])
def admin_edit_script_page(filename):
    p = os.path.join(BOTS_DIR, filename)
    content = ""
    if os.path.exists(p):
        with open(p, "r", encoding="utf-8") as f:
            content = f.read()
    return render_template_string(SCRIPT_EDITOR_HTML, filename=filename, content=content, glass_css=GLASS_CSS)

@app.route("/admin/scripts/save", methods=["POST"])
@check_role([ROLE_ADMIN])
def admin_save_script():
    filename = request.form.get("filename")
    content = request.form.get("content")
    p = os.path.join(BOTS_DIR, filename)
    with open(p, "w", encoding="utf-8") as f:
        f.write(content)
    flash("Fichier modifié avec succès.")
    return redirect(url_for('admin_edit_script_page', filename=filename))

@app.route("/admin/scripts/toggle", methods=["POST"])
@check_role([ROLE_ADMIN])
def admin_toggle_script():
    filename = request.form.get("filename")
    current_val = int(request.form.get("current_val", 1))
    new_val = 0 if current_val == 1 else 1
    db = get_db()
    db.execute("INSERT OR REPLACE INTO script_config (filename, is_active) VALUES (?, ?)", (filename, new_val))
    db.commit()
    return redirect(url_for('admin_scripts'))

@app.route("/admin/scripts/run", methods=["POST"])
@check_role([ROLE_ADMIN])
def admin_run_script_direct():
    filename = request.form.get("filename")
    p = os.path.join(BOTS_DIR, filename)
    if os.path.exists(p):
        if launch_script_core(filename, p, user_name=f"{session.get('username')} (Zone Admin)"):
            flash("Script initié en arrière-plan.")
        else:
            flash("Erreur : Un processus est déjà actif.")
    return redirect(url_for('index'))

@app.route("/admin/schedules", methods=["GET", "POST"])
@check_role([ROLE_ADMIN])
def admin_schedules():
    db = get_db()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            name = request.form.get("script_name")
            freq = request.form.get("frequency")
            val = int(request.form.get("time_value", 30))
            next_time = datetime.now() + timedelta(minutes=1)
            db.execute("INSERT INTO schedules (script_name, frequency, time_value, last_run, next_run) VALUES (?, ?, ?, 'Jamais', ?)",
                       (name, freq, val, next_time.strftime("%Y-%m-%d %H:%M:%S")))
            db.commit()
        elif action == "delete":
            db.execute("DELETE FROM schedules WHERE id = ?", (request.form.get("id"),))
            db.commit()
        return redirect(url_for('admin_schedules'))

    schedules = db.execute("SELECT * FROM schedules").fetchall()
    return render_template_string(SCHEDULES_HTML, schedules=schedules, scripts=get_all_bots(), glass_css=GLASS_CSS)

@app.route("/admin/stats")
@check_role([ROLE_ADMIN])
def admin_stats():
    db = get_db()
    total_users = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    total_logs = db.execute("SELECT COUNT(*) FROM logs").fetchone()[0]
    script_counts = db.execute("SELECT script, COUNT(*) as cnt FROM logs GROUP BY script ORDER BY cnt DESC").fetchall()
    return render_template_string(STATS_HTML, total_users=total_users, total_logs=total_logs, script_counts=script_counts, glass_css=GLASS_CSS)

# --- ROUTES SECURITÉ (CAPTCHA) ET NETTOYAGE LOGS ---

@app.route("/security-gate")
def security_gate():
    return render_template_string(GATE_HTML, site_key=RECAPTCHA_SITE_KEY, glass_css=GLASS_CSS)

@app.route("/verify-gate", methods=["POST"])
def verify_gate():
    captcha_response = request.form.get('g-recaptcha-response')
    if verify_recaptcha(captcha_response):
        session['captcha_passed'] = True
        return redirect(url_for('index'))
    flash("Veuillez valider le captcha correctement.")
    return redirect(url_for('security_gate'))

@app.route('/admin/clean_logs_manual', methods=['POST'])
@check_role([ROLE_ADMIN])
def clean_logs_manual():
    try:
        days = request.form.get('days', type=int)
        if days is not None and days >= 0:
            limit_date = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
            db = get_db()
            db.execute("DELETE FROM logs WHERE date < ?", (limit_date,))
            db.commit()
            flash(f"Logs plus anciens que {days} jours supprimés avec succès.")
        else:
            flash("Veuillez entrer un nombre de jours valide.")
    except Exception as e:
        flash(f"Erreur lors de la suppression : {str(e)}")
    return redirect(url_for('settings'))

# --- ROUTES STANDARDS ---

@app.route("/set_lang/<code>")
def set_language(code):
    if code not in ['fr', 'en']: code = 'fr'
    session['lang'] = code
    if 'user_id' in session:
        try:
            db = get_db()
            db.execute("UPDATE users SET lang = ? WHERE github_id = ?", (code, session['user_id']))
            db.commit()
        except: pass
    return redirect(request.referrer or url_for('index'))

@app.route("/admin/login")
def manual_login_page():
    return render_template_string(LOGIN_MANUAL_HTML, glass_css=GLASS_CSS)

@app.route("/admin/login_post", methods=['POST'])
def manual_login_post():
    username = request.form.get('username')
    password = request.form.get('password')
    if username == MANUAL_LOGIN_ID and password == MANUAL_LOGIN_PASS:
        db = get_db()
        existing_lang = db.execute("SELECT lang FROM users WHERE github_id = ?", (MANUAL_ADMIN_ID,)).fetchone()
        current_lang = existing_lang['lang'] if existing_lang else 'fr'
        db.execute("""INSERT OR REPLACE INTO users (github_id, username, avatar, role, is_banned, lang)
                        VALUES (?, ?, ?, ?, 0, ?)""",
                     (MANUAL_ADMIN_ID, MANUAL_ADMIN_USERNAME, "https://ui-avatars.com/api/?name=Admin+Bot&background=ff0000&color=fff", ROLE_ADMIN, current_lang))
        db.commit()

        session.permanent = True
        session['user_id'] = MANUAL_ADMIN_ID
        session['username'] = MANUAL_ADMIN_USERNAME
        session['lang'] = current_lang
        flash("Connexion Admin Manuelle Réussie.")
        return redirect(url_for('index'))
    flash(get_text('error_manual_login'))
    return redirect(url_for('manual_login_page'))

@app.route("/login")
def login_github():
    github_auth_url = (f"https://github.com/login/oauth/authorize?client_id={GITHUB_CLIENT_ID}"
                       f"&redirect_uri={GITHUB_REDIRECT_URI}&scope=user:email")
    return redirect(github_auth_url)

@app.route("/callback")
@csrf.exempt
def callback():
    code = request.args.get('code')
    if not code:
        return redirect(url_for('index'))
    data = {'client_id': GITHUB_CLIENT_ID, 'client_secret': GITHUB_CLIENT_SECRET, 'code': code}
    try:
        r = requests.post('https://github.com/login/oauth/access_token', data=data, headers={'Accept': 'application/json'})
        token_data = r.json()
        if 'access_token' not in token_data: return redirect(url_for('index'))

        r_user = requests.get(f"{GITHUB_API_BASE_URL}/user", headers={'Authorization': f"Bearer {token_data['access_token']}"})
        user_info = r_user.json()
        github_id, username = str(user_info['id']), user_info['login']
        avatar_url = user_info.get('avatar_url', 'https://avatars.githubusercontent.com/u/0?v=4')

        db = get_db()
        existing_user = db.execute("SELECT * FROM users WHERE github_id = ?", (github_id,)).fetchone()
        role, lang = ROLE_NONE, 'fr'
        if username == "janus": role = ROLE_ADMIN

        if existing_user:
            role, lang = existing_user['role'], existing_user['lang'] or 'fr'
            if existing_user['is_banned']: return redirect(url_for('index'))

        db.execute("""INSERT OR REPLACE INTO users (github_id, username, avatar, role, is_banned, lang, ban_reason)
                      VALUES (?, ?, ?, ?, COALESCE((SELECT is_banned FROM users WHERE github_id=?), 0), ?,
                              (SELECT ban_reason FROM users WHERE github_id=?))""",
                   (github_id, username, avatar_url, role, github_id, lang, github_id))
        db.commit()

        session['user_id'], session['username'], session['lang'] = github_id, username, lang
        return redirect(url_for('index'))
    except:
        return redirect(url_for('index'))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route("/account")
@login_required
def account():
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE github_id = ?", (session['user_id'],)).fetchone()
    return render_template_string(ACCOUNT_HTML, user=user, glass_css=GLASS_CSS)

@app.route("/admin/users")
@check_role([ROLE_ADMIN])
def admin_users():
    db = get_db()
    users = db.execute("SELECT * FROM users ORDER BY is_banned ASC, role DESC").fetchall()
    return render_template_string(ADMIN_USERS_HTML, users=users, glass_css=GLASS_CSS)

@app.route("/admin/update_user", methods=['POST'])
@check_role([ROLE_ADMIN])
def admin_update_user():
    target_id = request.form.get('github_id')
    new_role = request.form.get('new_role')
    action = request.form.get('action')
    reason = request.form.get('reason')

    if target_id == session['user_id']: return redirect(url_for('admin_users'))
    db = get_db()
    if action == 'ban':
        db.execute("UPDATE users SET is_banned = 1, ban_reason = ? WHERE github_id = ?", (reason, target_id))
    elif action == 'update' and new_role in [ROLE_NONE, ROLE_COLLAB, ROLE_ADMIN]:
        db.execute("UPDATE users SET role = ? WHERE github_id = ?", (new_role, target_id,))
    db.commit()
    return redirect(url_for('admin_users'))

@app.route("/settings")
@check_role([ROLE_ADMIN])
def settings():
    db = get_db()
    row_lock = db.execute("SELECT value FROM settings WHERE key='lock_launch'").fetchone()
    locked = row_lock['value'] if row_lock else '0'
    row_cap = db.execute("SELECT value FROM settings WHERE key='captcha_enabled'").fetchone()
    captcha_enabled = row_cap['value'] if row_cap else '0'
    row_clean = db.execute("SELECT value FROM settings WHERE key='auto_clean_days'").fetchone()
    auto_clean_days = row_clean['value'] if row_clean else '0'
    return render_template_string(SETTINGS_HTML, locked=locked, captcha_enabled=captcha_enabled, auto_clean_days=auto_clean_days, glass_css=GLASS_CSS)

@app.route('/update_settings', methods=['POST'])
@check_role([ROLE_ADMIN])
def update_settings():
    db = get_db()
    db.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", ('lock_launch', request.form.get('lock_launch', '0')))
    db.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", ('captcha_enabled', request.form.get('captcha_enabled', '0')))
    db.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", ('auto_clean_days', request.form.get('auto_clean_days', '0')))
    db.commit()
    return redirect(url_for('settings'))

@app.route("/history")
def history():
    f_search = request.args.get('search')
    query = "SELECT * FROM logs WHERE 1=1"
    params = []
    if f_search: query += " AND message LIKE ?"; params.append(f"%{f_search}%")
    query += " ORDER BY id DESC LIMIT 500"
    db = get_db()
    rows = db.execute(query, params).fetchall()

    count_query = "SELECT COUNT(*) FROM logs WHERE 1=1"
    if f_search: count_query += " AND message LIKE ?"
    total_count = db.execute(count_query, params).fetchone()[0]

    return render_template_string(HISTORY_HTML, rows=rows, request=request, glass_css=GLASS_CSS, total_count=total_count)


def _get_logs_for_export(f_search=None):
    """Retourne tous les logs (sans limite) pour l'export, avec filtre optionnel."""
    query = "SELECT id, date, script, message FROM logs WHERE 1=1"
    params = []
    if f_search:
        query += " AND message LIKE ?"
        params.append("%" + f_search + "%")
    query += " ORDER BY id DESC"
    db = get_db()
    return db.execute(query, params).fetchall()


@app.route("/history/export/excel")
def export_logs_excel():
    f_search = request.args.get('search')
    rows = _get_logs_for_export(f_search)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logs BotJanus"

    # Styles entête
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0B132B", end_color="0B132B", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = ["ID", "Date", "Script", "Message"]
    col_widths = [8, 22, 25, 80]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # Couleurs alternées pour les lignes
    fill_light = PatternFill(start_color="EAF4FB", end_color="EAF4FB", fill_type="solid")
    fill_dark = PatternFill(start_color="D0E8F2", end_color="D0E8F2", fill_type="solid")

    for row_idx, row in enumerate(rows, start=2):
        fill = fill_light if row_idx % 2 == 0 else fill_dark
        values = [row[0], row[1], row[2], row[3]]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 4))

    # Figer la première ligne
    ws.freeze_panes = "A2"

    # Auto-filtre
    ws.auto_filter.ref = "A1:D1"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "logs_botjanus_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=" + filename}
    )


@app.route("/history/export/csv")
def export_logs_csv():
    f_search = request.args.get('search')
    rows = _get_logs_for_export(f_search)

    si = io.StringIO()
    cw = csv.writer(si, delimiter=';')
    cw.writerow(['ID', 'Date', 'Script', 'Message'])
    for row in rows:
        cw.writerow([row[0], row[1], row[2], row[3]])

    filename = "logs_botjanus_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".csv"
    return Response(
        si.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=" + filename}
    )

@app.route("/")
def index():
    db = get_db()
    row = db.execute("SELECT value FROM settings WHERE key='lock_launch'").fetchone()
    locked = row['value'] if row else '0'
    user_role = ROLE_NONE
    if 'user_id' in session:
        u = db.execute("SELECT role FROM users WHERE github_id=?", (session['user_id'],)).fetchone()
        if u: user_role = u['role']

    # Filtrage dynamique des scripts selon l'autorisation de l'admin
    all_files = get_all_bots()
    available_scripts = []
    for f in all_files:
        is_act = get_script_status(f)
        if user_role == ROLE_ADMIN or is_act == 1:
            available_scripts.append({"filename": f, "is_active": is_act})

    return render_template_string(DASHBOARD_HTML, running=status["running"], script_name=status.get("script_name") or "Inactif",
                                  logs=status["live_output"][-50:], locked=locked, role=user_role, glass_css=GLASS_CSS,
                                  available_scripts=available_scripts)

@app.route("/api/live_logs")
def live_logs():
    return "\n".join([f"<div>{line}</div>" for line in status["live_output"][-100:]])

@app.route("/start", methods=["POST"])
@check_role([ROLE_COLLAB, ROLE_ADMIN])
def start_script():
    db = get_db()
    locked_val = db.execute("SELECT value FROM settings WHERE key='lock_launch'").fetchone()['value']
    user = db.execute("SELECT role FROM users WHERE github_id=?", (session['user_id'],)).fetchone()

    if locked_val == '1' and user['role'] != ROLE_ADMIN:
        flash("Verrouillé par l'Admin.")
        return redirect(url_for('index'))

    if not status["running"]:
        choice = request.form.get("choice")
        p = os.path.join(BOTS_DIR, choice)

        if user['role'] != ROLE_ADMIN and get_script_status(choice) == 0:
            flash("Ce script a été désactivé par l'administrateur.")
            return redirect(url_for('index'))

        if os.path.exists(p):
            cmd_args = []
            if "portal.py" in choice or choice == "Portail":
                arg_lang = request.form.get("arg_lang")
                arg_cat = request.form.get("arg_cat")
                arg_portal = request.form.get("arg_portal")
                if not (arg_lang and arg_cat and arg_portal): return redirect(url_for('index'))
                cmd_args = ["--lang", arg_lang, "--cat", arg_cat, "--portal", arg_portal]

            launch_script_core(choice, p, args=cmd_args, user_name=session.get('username'))
    return redirect(url_for("index"))

@app.route("/stop", methods=["POST"])
@check_role([ROLE_COLLAB, ROLE_ADMIN])
def stop_script():
    if status["running"] and status["process"]:
        try: os.kill(status["process"].pid, signal.SIGTERM)
        except: pass
        status["running"], status["script_name"], status["process"] = False, None, None
        log_to_db("SYSTEM", f"Arrêt par {session.get('username')}")
    return redirect(url_for("index"))

def read_output(process, script_name):
    for line in iter(process.stdout.readline, ''):
        cleaned = line.strip()
        if cleaned:
            status["last_activity"] = datetime.now()
            status["live_output"].append(cleaned)
            log_to_db(script_name, cleaned)
    process.stdout.close()
    status["running"] = False

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)